import os
import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton,
    QFileDialog, QTextEdit, QLabel, QMessageBox, QComboBox
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET

# Helper Functions
def get_week_start(dt):
    if dt.weekday() == 6:
        week_start = dt
    else:
        week_start = dt - timedelta(days=(dt.weekday() + 1))
    return week_start.strftime("%Y-%m-%d")

def get_month_key(dt):
    """Return the year-month key (e.g., '2025-04') for a datetime object."""
    return dt.strftime("%Y-%m")

def choose_timestamp(ts_list):
    try:
        dt_list = [datetime.strptime(ts, "%m/%d/%y %H:%M:%S") for ts in ts_list]
        earliest = min(dt_list)
        return earliest.strftime("%m/%d/%y %H:%M:%S")
    except Exception:
        return ts_list[0] if ts_list else ""

# XML Parsing and Data Processing
def parse_xml(xml_file):
    try:
        tree = ET.parse(xml_file)
        return tree.getroot()
    except Exception as e:
        return None, f"Error parsing {xml_file}: {e}"

def extract_monitor_sources(xml_root):
    """
    Extract unique monitor sources from the XML by scanning <details> elements.
    Returns a list of unique monitor source strings.
    """
    monitor_sources = set()
    entries = xml_root.findall('.//entry')
    for entry in entries:
        details_elem = entry.find('details')
        if details_elem is None or not details_elem.text:
            continue
        details = details_elem.text.lower()
        # Split details into words and look for potential monitor sources
        # This is a basic heuristic; adjust based on XML structure
        words = details.split()
        for word in words:
            # Filter out common words, numbers, and short strings
            if len(word) > 3 and not word.isdigit() and word not in ["required", "weekly", "monthly", "test", "sent", "received"]:
                monitor_sources.add(word)
    return sorted(list(monitor_sources))

def process_entries(xml_root, monitor_sources_order):
    """
    Process each <entry> element in the XML using dynamically provided monitor sources.
    """
    weekly_data = {}
    monthly_data = {}
    weekly_weeks = set()
    monthly_weeks = set()

    entries = xml_root.findall('.//entry')
    for entry in entries:
        details_elem = entry.find('details')
        date_elem = entry.find('date')
        type_elem = entry.find('type')
        if details_elem is None or date_elem is None or type_elem is None:
            continue

        details = details_elem.text or ""
        date_str = date_elem.text.strip() if date_elem.text else ""
        alert_type_text = type_elem.text.strip() if type_elem.text else ""
        details_lower = details.lower()

        is_weekly = ("rwt" in details_lower) or ("required weekly test" in details_lower)
        is_monthly = ("rmt" in details_lower) or ("required monthly test" in details_lower)
        if not (is_weekly or is_monthly):
            continue

        monitor_lp_source = None
        for src in monitor_sources_order:
            if src.lower() in details_lower:
                monitor_lp_source = src
                break

        if not monitor_lp_source:
            continue

        try:
            dt = datetime.strptime(date_str, "%m/%d/%y %H:%M:%S")
        except ValueError:
            continue

        week_start = get_week_start(dt)
        if is_weekly:
            data_dict = weekly_data
            weekly_weeks.add(week_start)
        else:
            data_dict = monthly_data
            monthly_weeks.add(week_start)

        alert_type = alert_type_text.lower()
        if alert_type == "sent":
            direction = "Sent"
        elif alert_type == "received":
            direction = "Received"
        else:
            continue

        key = (monitor_lp_source, direction)
        if key not in data_dict:
            data_dict[key] = {}
        if week_start not in data_dict[key]:
            data_dict[key][week_start] = []
        data_dict[key][week_start].append(date_str)

    return weekly_data, sorted(weekly_weeks), monthly_data, sorted(monthly_weeks)

# Cell Formatting
red_fill = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")

# Table Append Functions
def append_weekly_table(ws, start_row, table_title, data_dict, week_columns, direction_filter):
    header = ["Monitor/LP"] + week_columns + ["Notes"]
    num_columns = len(header)
    
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_columns)
    title_cell = ws.cell(row=start_row, column=1, value=table_title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    header_row = start_row + 1
    for col, head in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col, value=head)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    current_row = header_row + 1
    for (monitor_lp, direction), weeks_dict in sorted(data_dict.items(), key=lambda x: (x[0][0], x[0][1])):
        if direction != direction_filter:
            continue
        row_data = [monitor_lp]
        for week in week_columns:
            ts_list = weeks_dict.get(week, [])
            cell_value = choose_timestamp(ts_list) if ts_list else "Not parsed"
            row_data.append(cell_value)
        row_data.append("")
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if isinstance(value, str) and "not parsed" in value.lower():
                cell.fill = red_fill
        current_row += 1
    
    return current_row + 1

def aggregate_monthly_single(monthly_data):
    aggregated = {}
    for key, week_dict in monthly_data.items():
        monitor_lp, direction = key
        for week, ts_list in week_dict.items():
            if not ts_list:
                continue
            chosen = choose_timestamp(ts_list)
            try:
                dt = datetime.strptime(chosen, "%m/%d/%y %H:%M:%S")
            except Exception:
                continue
            if direction not in aggregated or dt < aggregated[direction][0]:
                aggregated[direction] = (dt, chosen, monitor_lp)
    return aggregated

def append_monthly_table_new(ws, start_row, table_title, monthly_data, total_cols):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    title_cell = ws.cell(row=start_row, column=1, value=table_title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="Monitor/LP").font = Font(bold=True)
    ws.cell(row=header_row, column=2, value="Alert Type").font = Font(bold=True)
    ws.cell(row=header_row, column=3, value="Timestamp").font = Font(bold=True)
    ws.merge_cells(start_row=header_row, start_column=4, end_row=header_row, end_column=total_cols)
    ws.cell(row=header_row, column=4, value="Notes").font = Font(bold=True)
    for col in range(1, total_cols + 1):
        ws.cell(row=header_row, column=col).alignment = Alignment(horizontal="center")
    
    aggregated = aggregate_monthly_single(monthly_data)
    
    data_rows = []
    for direction in ["Received", "Sent"]:
        if direction in aggregated:
            dt, ts, monitor_lp = aggregated[direction]
            data_rows.append([monitor_lp, "RMT", ts, ""])
        else:
            data_rows.append(["", "RMT", "Not parsed", ""])
    
    current_row = header_row + 1
    for row_data in data_rows:
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col, value=row_data[col-1])
            if isinstance(row_data[col-1], str) and "not parsed" in row_data[col-1].lower():
                cell.fill = red_fill
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=total_cols)
        ws.cell(row=current_row, column=4, value=row_data[3])
        current_row += 1
    
    ws.cell(row=current_row, column=1, value="Within 1 Hour?").font = Font(bold=True)
    if "Received" in aggregated and "Sent" in aggregated:
        rcv_dt, _, _ = aggregated["Received"]
        sent_dt, _, _ = aggregated["Sent"]
        diff = (sent_dt - rcv_dt).total_seconds()
        within_val = "Yes" if 0 <= diff <= 3600 else "No"
    else:
        within_val = "N/A"
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=total_cols)
    ws.cell(row=current_row, column=2, value=within_val)
    current_row += 1
    
    return current_row + 1

def append_weekly_eas_review(ws, start_row, weeks):
    total_cols = len(weeks) + 2
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    title_cell = ws.cell(row=start_row, column=1, value="Weekly EAS Review")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="Week").font = Font(bold=True)
    ws.cell(row=header_row, column=2, value="Initials").font = Font(bold=True)
    ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=total_cols)
    ws.cell(row=header_row, column=3, value="Notes").font = Font(bold=True)
    for col in range(1, total_cols+1):
        ws.cell(row=header_row, column=col).alignment = Alignment(horizontal="center")

    current_row = header_row + 1
    for week in weeks:
        ws.cell(row=current_row, column=1, value=week)
        ws.cell(row=current_row, column=2, value="")
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=total_cols)
        ws.cell(row=current_row, column=3, value="")
        current_row += 1
    return current_row + 2

def auto_adjust_column_widths(ws):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# GUI Application
class EASParserApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("EAS XML Parser")
        self.setGeometry(100, 100, 800, 600)
        self.xml_files = []
        self.output_dir = ""
        self.monitor_sources = []
        self.selected_monitor_sources = []
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Sample XML for Monitor Sources
        self.sample_label = QLabel("No sample XML file selected for monitor sources")
        layout.addWidget(self.sample_label)
        select_sample_btn = QPushButton("Select Sample XML to Detect Monitor Sources")
        select_sample_btn.clicked.connect(self.select_sample_xml)
        layout.addWidget(select_sample_btn)

        # Monitor Sources Selection
        self.monitor_label = QLabel("Monitor Sources: Not detected")
        layout.addWidget(self.monitor_label)
        self.monitor_combo = QComboBox()
        self.monitor_combo.addItem("Select Monitor Sources")
        self.monitor_combo.currentIndexChanged.connect(self.update_monitor_sources)
        layout.addWidget(self.monitor_combo)

        # XML Files Selection
        self.xml_label = QLabel("No XML files selected")
        layout.addWidget(self.xml_label)
        select_xml_btn = QPushButton("Select XML Files to Parse")
        select_xml_btn.clicked.connect(self.select_xml_files)
        layout.addWidget(select_xml_btn)

        # Output Directory Selection
        self.output_label = QLabel("No output directory selected")
        layout.addWidget(self.output_label)
        select_output_btn = QPushButton("Select Output Directory")
        select_output_btn.clicked.connect(self.select_output_dir)
        layout.addWidget(select_output_btn)

        # Parse Button
        parse_btn = QPushButton("Parse and Generate Excel Reports")
        parse_btn.clicked.connect(self.parse_files)
        layout.addWidget(parse_btn)

        # Log Area
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        layout.addWidget(self.log_area)

    def log(self, message):
        self.log_area.append(message)

    def select_sample_xml(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "Select Sample XML File", "", "XML Files (*.xml)"
        )
        if file:
            self.sample_label.setText(f"Sample XML: {file}")
            self.log(f"Selected sample XML: {file}")
            xml_root = parse_xml(file)
            if xml_root is None:
                self.log(f"Failed to parse sample XML: {file}")
                return
            self.monitor_sources = extract_monitor_sources(xml_root)
            if not self.monitor_sources:
                self.log("No monitor sources detected in the sample XML.")
                self.monitor_label.setText("Monitor Sources: None detected")
                return
            self.monitor_combo.clear()
            self.monitor_combo.addItem("Select Monitor Sources")
            self.monitor_combo.addItems(self.monitor_sources)
            self.monitor_label.setText(f"Monitor Sources: {len(self.monitor_sources)} detected")
            self.log(f"Detected monitor sources: {', '.join(self.monitor_sources)}")
        else:
            self.sample_label.setText("No sample XML file selected for monitor sources")

    def update_monitor_sources(self):
        selected = self.monitor_combo.currentText()
        if selected == "Select Monitor Sources":
            self.selected_monitor_sources = self.monitor_sources
        else:
            self.selected_monitor_sources = [selected]
        self.log(f"Selected monitor sources for processing: {', '.join(self.selected_monitor_sources)}")

    def select_xml_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select XML Files", "", "XML Files (*.xml)"
        )
        if files:
            self.xml_files = files
            self.xml_label.setText(f"Selected {len(files)} XML file(s)")
            self.log(f"Selected XML files: {', '.join(files)}")
        else:
            self.xml_label.setText("No XML files selected")

    def select_output_dir(self):
        directory = QFileDialog.getExistingDirectory(
            self, "Select Output Directory"
        )
        if directory:
            self.output_dir = directory
            self.output_label.setText(f"Output directory: {directory}")
            self.log(f"Output directory set to: {directory}")
        else:
            self.output_label.setText("No output directory selected")

    def parse_files(self):
        if not self.xml_files:
            QMessageBox.warning(self, "Warning", "Please select at least one XML file to parse.")
            return
        if not self.output_dir:
            QMessageBox.warning(self, "Warning", "Please select an output directory.")
            return
        if not self.monitor_sources:
            QMessageBox.warning(self, "Warning", "Please select a sample XML file to detect monitor sources.")
            return
        if not self.selected_monitor_sources:
            QMessageBox.warning(self, "Warning", "Please select monitor sources to use for parsing.")
            return

        self.log("Starting processing...")
        all_weekly_data = {}
        all_monthly_data = {}
        all_weekly_weeks = set()
        all_monthly_weeks = set()
        month_key = None

        for xml_file in self.xml_files:
            self.log(f"Processing {xml_file}...")
            xml_root = parse_xml(xml_file)
            if xml_root is None:
                self.log(f"Failed to parse {xml_file}")
                continue

            weekly_data, weekly_weeks, monthly_data, monthly_weeks = process_entries(xml_root, self.selected_monitor_sources)
            
            for week in weekly_weeks:
                dt = datetime.strptime(week, "%Y-%m-%d")
                if month_key is None:
                    month_key = get_month_key(dt)
                elif get_month_key(dt) != month_key:
                    self.log(f"Warning: XML file {xml_file} contains data from a different month ({get_month_key(dt)}). Skipping.")
                    continue

            for key, weeks_dict in weekly_data.items():
                if key not in all_weekly_data:
                    all_weekly_data[key] = {}
                for week, ts_list in weeks_dict.items():
                    if week not in all_weekly_data[key]:
                        all_weekly_data[key][week] = []
                    all_weekly_data[key][week].extend(ts_list)
                    all_weekly_weeks.add(week)

            for key, weeks_dict in monthly_data.items():
                if key not in all_monthly_data:
                    all_monthly_data[key] = {}
                for week, ts_list in weeks_dict.items():
                    if week not in all_monthly_data[key]:
                        all_monthly_data[key][week] = []
                    all_monthly_data[key][week].extend(ts_list)
                    all_monthly_weeks.add(week)

        if not month_key:
            self.log("No valid data found to determine the month.")
            return

        output_file = os.path.join(self.output_dir, f"EAS_{month_key}.xlsx")
        
        if os.path.exists(output_file):
            try:
                wb = load_workbook(output_file)
                self.log(f"Updating existing file: {output_file}")
            except Exception as e:
                self.log(f"Error loading existing file {output_file}: {e}. Creating new file.")
                wb = Workbook()
        else:
            self.log(f"Creating new file: {output_file}")
            wb = Workbook()

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        if "EAS Alerts" not in wb.sheetnames:
            ws = wb.create_sheet("EAS Alerts")
        else:
            ws = wb["EAS Alerts"]

        ws.delete_rows(1, ws.max_row)

        weeks = sorted(all_weekly_weeks)
        total_cols = len(weeks) + 2

        current_row = 1
        current_row = append_weekly_table(ws, current_row, "Required Weekly Tests (Received)", all_weekly_data, weeks, "Received")
        current_row = append_weekly_table(ws, current_row, "Required Weekly Tests (Sent)", all_weekly_data, weeks, "Sent")
        current_row = append_monthly_table_new(ws, current_row, "Required Monthly Tests", all_monthly_data, total_cols)
        current_row = append_weekly_eas_review(ws, current_row, weeks)

        auto_adjust_column_widths(ws)

        try:
            wb.save(output_file)
            self.log(f"Data successfully saved to {output_file}")
        except Exception as e:
            self.log(f"Error saving Excel file: {e}")

        self.log("Processing complete.")

def main():
    app = QApplication(sys.argv)
    window = EASParserApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()